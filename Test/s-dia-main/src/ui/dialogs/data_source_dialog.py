from PySide6.QtWidgets import (QDialog, QMessageBox, QLineEdit, QComboBox, QPushButton, 
                              QTableWidget, QTableWidgetItem, QCheckBox, QWidget, QHBoxLayout, QVBoxLayout, QLabel, QMenu, QFileDialog, QHeaderView, QStyledItemDelegate, QApplication, QStyleOptionButton, QStyle)
from PySide6.QtGui import QStandardItemModel, QStandardItem, QAction, QIcon, QMovie, QColor
from PySide6.QtCore import Qt, QEvent, QTimer, QSize
from PySide6.QtUiTools import QUiLoader
from ups.dask.table import read_csv, read_parquet
from ui.project_schema import Schema
from utils.enum_utils import cnv_str_to_datatype, dataTypeToView, viewToDataType, deIdentificationAttributeToView, viewToDeIdentificationAttribute

from defined import DeIdentificationAttribute, DataType
import pandas as pd
import os
import sys
import logging
import platform
import uuid
import chardet
import json
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from utils.string_utils import get_file_extension
from ..common.context_menu import CustomContextMenuFilter

class ClickCheckBoxDelegate(QStyledItemDelegate):
    def createEditor(self, parent, option, index):
        editor = super().createEditor(parent, option, index)
        if isinstance(editor, QLineEdit):
            editor.setFixedHeight(20)  # 높이 20px로 고정
            editor.setContentsMargins(0, 0, 0, 0)  # 컨텐츠 마진 제거
            editor.setStyleSheet("""
                QLineEdit {
                    background-color: white;
                    border: 1px solid #339af0;
                    border-radius: 4px;
                    padding: 0px 2px;
                    margin: 0px;
                    color: #495057;
                    min-height: 20px;
                    max-height: 20px;
                    height: 20px;
                }
                QLineEdit:focus {
                    border: 1px solid #339af0;
                    outline: none;
                }
            """)
            # 텍스트 정렬을 중앙으로 설정
            editor.setAlignment(Qt.AlignCenter)
        return editor

    def setEditorData(self, editor, index):
        if isinstance(editor, QCheckBox):
            value = index.data(Qt.EditRole)
            editor.setChecked(bool(value))
        else:
            super().setEditorData(editor, index)

    def setModelData(self, editor, model, index):
        if isinstance(editor, QCheckBox):
            model.setData(index, editor.isChecked(), Qt.EditRole)
        else:
            super().setModelData(editor, model, index)

    def updateEditorGeometry(self, editor, option, index):
        if isinstance(editor, QLineEdit):
            # 편집기의 위치와 크기를 셀에 맞게 조정
            rect = option.rect
            editor.setGeometry(rect.x(), rect.y() + (rect.height() - 20) // 2, rect.width(), 20)
        else:
            super().updateEditorGeometry(editor, option, index)

class DataSourceDialog(QDialog):
    def __init__(self, parent=None, data=None):
        super().__init__(parent)

        self.data = data
        self.schemas = {}
        self.logger = logging.getLogger(__name__)
        self.context_menu_filter = CustomContextMenuFilter()
        
        # 프로젝트 데이터와 데이터 소스 목록 저장
        if parent and hasattr(parent, 'project_data'):
            self.project_data = parent.project_data
            self.data_sources = parent._get_current_data_sources()
        else:
            self.project_data = None
            self.data_sources = []
        
        # 다이얼로그의 초기 크기 설정
        self.setMinimumSize(800, 627)  # 최소 너비 800px, 최소 높이 600px
        self.resize(800, 627)  # 초기 크기를 최소 크기로 설정
        
        self.loading_overlay = None  # 오버레이 위젯 참조용
        
        self.setup_ui()
        
        # 크기 변경 이벤트 처리
        self.installEventFilter(self)

    def setup_ui(self):
        try:
            # UI 파일 로드
            loader = QUiLoader()
            self.ui = loader.load("src/ui/dialogs/data_source_dialog.ui")
            if not self.ui:
                self.logger.error("데이터 소스 다이얼로그 UI 파일을 로드할 수 없습니다.")
                return

            self.edit_flag = False
            
            # 위젯 참조
            self.name_input = self.ui.findChild(QLineEdit, "name_input")
            self.data_path_input = self.ui.findChild(QLineEdit, "data_path_input")
            self.charset_input = self.ui.findChild(QComboBox, "charset_input")
            self.separator_input = self.ui.findChild(QComboBox, "separator_input")
            self.has_header_input = self.ui.findChild(QComboBox, "has_header_input")
            self.save_button = self.ui.findChild(QPushButton, "save_button")
            self.cancel_button = self.ui.findChild(QPushButton, "cancel_button")
            self.fs_explorer = self.ui.findChild(QPushButton, "fs_explorer")
            self.schema_table = self.ui.findChild(QTableWidget, "schema_table")
            self.title_label = self.ui.findChild(QLabel, "title_label")
            self.schema_edit_btn = self.ui.findChild(QPushButton, "schema_edit_btn")
            self.schema_upload_btn = self.ui.findChild(QPushButton, "schema_upload_btn")
            self.schema_download_btn = self.ui.findChild(QPushButton, "schema_download_btn")
            self.csv_container = self.ui.findChild(QWidget, "csv_container")
            self.message_container = self.ui.findChild(QWidget, "message_container")
            self.message_text = self.ui.findChild(QLabel, "message_text")

            # 스키마 편집 버튼 아이콘 설정
            self.schema_edit_btn.setIcon(QIcon("src/ui/resources/images/icons8-lock-96.png"))
            self.schema_edit_btn.setIconSize(QSize(16, 16))
            self.schema_edit_btn.setText("")
            self.schema_edit_btn.setToolTip("스키마 편집 잠김")

            # 스키마 업로드/다운로드 버튼 아이콘 설정
            self.schema_upload_btn.setIcon(QIcon("src/ui/resources/images/upload-bold-arrow-icon.png"))
            self.schema_upload_btn.setIconSize(QSize(24, 24))
            self.schema_upload_btn.setText("")
            self.schema_upload_btn.setToolTip("스키마 Excel 파일 적용")
            
            self.schema_download_btn.setIcon(QIcon("src/ui/resources/images/download-bold-arrow-icon.png"))
            self.schema_download_btn.setIconSize(QSize(24, 24))
            self.schema_download_btn.setText("")
            self.schema_download_btn.setToolTip("스키마 Excel 파일 생성")
            
            # 스키마 테이블 설정
            self.setup_schema_table([])

            # 데이터 타입과 비식별 속성 컬럼만 편집 가능하도록 설정
            self.schema_table.setEditTriggers(QTableWidget.CurrentChanged)  # 싱글 클릭으로 변경
            
            self.name_input.installEventFilter(self.context_menu_filter)
            self.data_path_input.installEventFilter(self.context_menu_filter)
            # self.schema_table.installEventFilter(self.context_menu_filter)  # 테이블 위젯에 컨텍스트 메뉴 필터 적용

            # 순서, 컬럼명, 값1, 값2, 값3 컬럼은 읽기 전용으로 설정
            for row in range(self.schema_table.rowCount()):
                for col in [0, 1, 4, 5, 6]:
                    item = self.schema_table.item(row, col)
                    if item:
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)

            self.message_container.hide()
            self.csv_container.hide()

            # 초기 위젯 크기 조절
            self.adjust_widget_sizes()
            
            # 콤보박스 스타일 적용
            for combo in [self.charset_input, self.separator_input, self.has_header_input]:
                view = combo.view()
                view.setItemAlignment(Qt.AlignCenter)

            # 수정 모드인 경우 기존 데이터 설정
            self.show_loading()
            QTimer.singleShot(0, self._after_show_setup)
            
            # 시그널 연결
            self.save_button.clicked.connect(self.accept)
            self.cancel_button.clicked.connect(self.reject)
            self.fs_explorer.clicked.connect(self.browse_data_path)
            self.charset_input.currentTextChanged.connect(self.load_data)
            self.separator_input.currentTextChanged.connect(self.load_data)
            self.has_header_input.currentTextChanged.connect(self.load_data)
            self.schema_edit_btn.clicked.connect(self.edit_schema)
            self.schema_download_btn.clicked.connect(self.download_schema_excel)
            self.schema_upload_btn.clicked.connect(self.upload_schema_excel)

            # 레이아웃 설정
            self.setLayout(self.ui.layout())
            
            self.schema_table.setSelectionMode(QTableWidget.SingleSelection)
            # self.schema_table.setFocusPolicy(Qt.NoFocus)
            # QTableView와 유사한 밝은 헤더/셀 스타일 적용
            self.schema_table.setStyleSheet("""
                QTableWidget {
                    background-color: #f8f9fa;
                    border: 1px solid #dee2e6;
                    border-radius: 4px;
                    gridline-color: #e9ecef;
                    color: #212529;
                }
                QTableWidget::item {
                    padding: 4px;
                    border-bottom: 1px solid #e9ecef;
                    color: #212529;
                    background-color: white;
                }
                QTableWidget::item:selected {
                    background-color: #e7f5ff;
                    color: #212529;
                }
                QTableWidget::item:focus {
                    outline: none;
                    border: none;
                    background: #e7f5ff;  /* 선택 배경만 남기고 테두리 제거 */
                }
                QHeaderView::section {
                    background-color: #f8f9fa;
                    color: #212529;
                    padding: 8px;
                    border: none;
                    border-right: 1px solid #dee2e6;
                    border-bottom: 1px solid #dee2e6;
                    font-weight: bold;
                }
                QScrollBar:vertical {
                    border: none;
                    background: #f8f9fa;
                    width: 10px;
                    margin: 0px;
                }
                QScrollBar::handle:vertical {
                    background: #dee2e6;
                    min-height: 20px;
                    border-radius: 5px;
                    min-height: 30px;
                }
                QScrollBar::handle:vertical:hover {
                    background: #adb5bd;
                }
                QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {
                    height: 0px;
                }
                QScrollBar::add-page:vertical, QScrollBar::sub-page:vertical {
                    background: none;
                }
                QScrollBar:horizontal {
                    border: none;
                    background: #f8f9fa;
                    height: 10px;
                    margin: 0px;
                }
                QScrollBar::handle:horizontal {
                    background: #dee2e6;
                    min-width: 20px;
                    border-radius: 5px;
                }
                QScrollBar::handle:horizontal:hover {
                    background: #adb5bd;
                }
                QScrollBar::add-line:horizontal, QScrollBar::sub-line:horizontal {
                    width: 0px;
                }
                QScrollBar::add-page:horizontal, QScrollBar::sub-page:horizontal {
                    background: none;
                }
            """)
        
            self.create_loading_overlay()
            
            # 스크롤바 컨텍스트 메뉴 비활성화
            self.schema_table.verticalScrollBar().setContextMenuPolicy(Qt.NoContextMenu)
            self.schema_table.horizontalScrollBar().setContextMenuPolicy(Qt.NoContextMenu)
            
        except Exception as e:
            self.logger.error(f"데이터 소스 다이얼로그 UI 설정 중 오류 발생: {str(e)}")
            import traceback
            self.logger.error(traceback.format_exc())

    def _after_show_setup(self):
        try:
            if self.data:
                data_path = self.data.get('data_path', '')
                self.title_label.setText("데이터 소스 수정")
                self.name_input.setText(self.data.get('name', ''))
                self.data_path_input.setText(self.data.get('data_path', ''))
                self.charset_input.setCurrentText(self.data.get('charset', 'UTF-8'))
                self.separator_input.setCurrentText(self.data.get('separator', ','))
                self.has_header_input.setCurrentText(self.data.get('has_header', '있음'))
                self.uuid = self.data.get('uuid', str(uuid.uuid4()))

                schemas = self.data.get('schemas', [])
                if schemas:
                    schema_objects = []
                    self.schemas.clear()
                    for schema in schemas:
                        schema_obj = Schema()
                        schema_obj.name = schema.get('name', '')
                        schema_obj.comment = schema.get('comment', '')
                        try:
                            schema_obj.data_type = DataType[schema.get('data_type', 'STRING')]
                        except (KeyError, ValueError):
                            schema_obj.data_type = DataType.STRING
                        try:
                            schema_obj.de_identi_attr = DeIdentificationAttribute[schema.get('de_identi_attr', 'NONE')]
                        except (KeyError, ValueError):
                            schema_obj.de_identi_attr = DeIdentificationAttribute.NONE
                        schema_obj.is_distribution = schema.get('is_distribution', False)
                        schema_obj.is_statistics = schema.get('is_statistics', False)
                        schema_objects.append(schema_obj)
                        self.schemas[schema_obj.name] = schema_obj

                    # 원본 데이터에서 샘플 값 추출
                    first_rows = None
                    if data_path and os.path.exists(data_path):
                        self.is_csv = True if self.data.get('data_type', 'csv').lower() == 'csv' else False   #data_path.lower().endswith('.csv')
                        try:
                            if self.is_csv:
                                header_line = 0 if self.data.get('has_header', '있음') == '있음' else None
                                sep = self.data.get('separator', ',')
                                if sep == 'TAB':
                                    sep = '\t'
                                df = pd.read_csv(data_path, encoding=self.data.get('charset', 'utf-8'), sep=sep, header=header_line)
                            else:
                                df = pd.read_parquet(data_path)
                            first_rows = df.head(3)
                        except Exception as e:
                            self.logger.error(f"샘플 값 추출 중 오류: {str(e)}")
                            first_rows = None
                    self.display_schema_data(schema_objects, first_rows)
            else:
                self.uuid = str(uuid.uuid4())
                self.title_label.setText("데이터 소스 추가")
        finally:
            self.hide_loading()
            # 원본 데이터 파일 체크 및 메시지
            data_path = self.data.get('data_path', '') if self.data else ''
            if data_path and not os.path.exists(data_path):
                self.message_text.setText("원본 데이타가 없습니다. 원본 데이터를 선택해주세요.")
                self.message_container.show()
                # QMessageBox.warning(self, "경고", "원본 데이타가 없습니다. 원본 데이터를 선택해주세요.")
            self.adjust_widget_sizes()

    def create_loading_overlay(self):
        """로딩 오버레이 위젯 생성"""
        self.loading_overlay = QWidget(self)
        self.loading_overlay.setStyleSheet("background: rgba(255,255,255,180); border-radius: 8px;")
        self.loading_overlay.setGeometry(0, 0, self.width(), self.height())
        self.loading_overlay.hide()

        layout = QVBoxLayout(self.loading_overlay)
        layout.setAlignment(Qt.AlignCenter)

        # 로딩 아이콘 (GIF)
        self.loading_movie = QMovie("src/ui/resources/images/loading.gif")  # 스피너 GIF 경로
        loading_label = QLabel()
        loading_label.setMovie(self.loading_movie)
        layout.addWidget(loading_label)

        # 텍스트
        text_label = QLabel("로딩 중...")
        text_label.setStyleSheet("font-size: 16px; color: #555;")
        layout.addWidget(text_label)

        # 다이얼로그 크기 변경 시 오버레이도 같이 조정
        self.resizeEvent = self._resize_loading_overlay

    def _resize_loading_overlay(self, event):
        if self.loading_overlay:
            self.loading_overlay.setGeometry(0, 0, self.width(), self.height())
        QDialog.resizeEvent(self, event)

    def setup_schema_table(self, data_list):
        """스키마 테이블 설정 (QTableWidget 버전)"""
        # 테이블 모델 생성
        self.schema_table.setRowCount(len(data_list))
        self.schema_table.setColumnCount(10)
        headers = ["순서", "컬럼명", "설명", "데이터 타입", "개인정보속성", "분포", "통계", "값1", "값2", "값3"]
        self.schema_table.setHorizontalHeaderLabels(headers)
        
        # 컬럼 너비 설정
        self.schema_table.setColumnWidth(0, 50)   # 순서
        self.schema_table.setColumnWidth(1, 100)  # 컬럼명
        self.schema_table.setColumnWidth(2, 150)  # 설명
        self.schema_table.setColumnWidth(3, 95)   # 데이터 타입
        self.schema_table.setColumnWidth(4, 95)   # 개인정보속성
        self.schema_table.setColumnWidth(5, 40)   # 분포
        self.schema_table.setColumnWidth(6, 40)   # 통계
        # self.schema_table.setColumnWidth(7, 100)   # 값1
        # self.schema_table.setColumnWidth(8, 100)   # 값2
        # self.schema_table.setColumnWidth(9, 100)   # 값3

        # 특정 컬럼의 크기 조절 비활성화
        header = self.schema_table.horizontalHeader()
        header.setSectionResizeMode(3, QHeaderView.Fixed)  # 데이터 타입
        header.setSectionResizeMode(4, QHeaderView.Fixed)  # 개인정보속성
        header.setSectionResizeMode(5, QHeaderView.Fixed)  # 분포
        header.setSectionResizeMode(6, QHeaderView.Fixed)  # 통계

        # 커스텀 델리게이트 설정
        self.schema_table.setItemDelegate(ClickCheckBoxDelegate())

        # 기타 속성 설정
        self.schema_table.setAlternatingRowColors(True)
        self.schema_table.setSelectionBehavior(QTableWidget.SelectRows)
        self.schema_table.setEditTriggers(QTableWidget.CurrentChanged)  # 싱글 클릭으로 변경
        self.schema_table.verticalHeader().setVisible(False)
        self.schema_table.horizontalHeader().setVisible(True)
        self.schema_table.setShowGrid(True)
        self.schema_table.horizontalHeader().setSectionsMovable(False)
        self.schema_table.setFixedSize(730, 161)

    combo_stylesheet = """
            QComboBox {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 2px 2px;
                color: #495057;
            }
            QComboBox:focus {
                border: 1px solid #339af0;
            }
            QComboBox:disabled, QComboBox:disabled:editable, QComboBox:disabled:!editable {
                border: 0px;
                background: transparent;
                color: #495057;
            }
            QComboBox::drop-down {
                subcontrol-origin: padding;
                subcontrol-position: center right;
                width: 30px;
                border: none;
            }
            QComboBox:disabled::drop-down {
                width: 0;
                border: none;
                background: transparent;
            }
            QComboBox:disabled::down-arrow {
                width: 0;
                height: 0;
                image: none;
                border: none;
                background: transparent;
            }      
            QComboBox::down-arrow {
                image: url(src/ui/resources/images/down-arrow-svgrepo-com.svg);
                width: 10px;
                height: 10px;
                margin-right: 1px;
            }
            QComboBox QAbstractItemView {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                outline: none;
                selection-background-color: #e9ecef;
                selection-color: #000;
            }
            QComboBox QAbstractItemView::item {
                padding: 2px 2px;
                color: #495057;
            }
            QComboBox QAbstractItemView::item:selected {
                background-color: #e9ecef;
                color: #000;
            }
        """
    
    def combo_box(self, items, value, width):
        widget = QWidget()
        layout = QHBoxLayout(widget)
        layout.setContentsMargins(0, 0, 0, 0)
        layout.setAlignment(Qt.AlignCenter)

        combo_box = QComboBox()
        combo_box.addItems(items)
        combo_box.setCurrentText(value)
        combo_box.setFixedHeight(20)  # 높이 고정
        combo_box.setFixedWidth(width)
        combo_box.setStyleSheet(self.combo_stylesheet)
        
        layout.addWidget(combo_box)
        return widget, combo_box

    def data_type_combo_box(self, data_type):
        return self.combo_box(["문자", "정수", "실수", "날짜", "일시"], str(data_type), 83)
    
    def deid_attr_combo_box(self, deid_attr_value):
        return self.combo_box(["", "ID", "QI", "NSA", "SA"], str(deid_attr_value), 83)

    def display_schema_data(self, schemas, first_rows=None):
        try:
            # 테이블 초기화
            self.schema_table.setRowCount(0)
            if self.data:
                self.is_csv = True if self.data.get('data_type', 'csv') == 'csv' else False
            else:
                file_path = self.data_path_input.text()
                self.is_csv = True if file_path.lower().endswith('.csv') else False

            # self.schemas.clear()

            # enumerate를 사용하여 idx와 Schema 객체를 동시에 가져옴
            self.edit_control = []
            checkbox_stylesheet = """
                QCheckBox {
                    background: transparent;
                }
                QCheckBox:disabled {
                    background-color: #e9ecef;
                }
                QCheckBox::indicator {
                    width: 16px;
                    height: 16px;
                    border: 1px solid #787878;
                    border-radius: 4px;
                    background-color: white;
                }
                QCheckBox::indicator:unchecked {
                    background-color: white;
                }
                QCheckBox::indicator:checked {
                    background-color: white;
                    image: url(src/ui/resources/images/check.png);
                }
            """ 

            for idx, schema_item in enumerate(schemas):
                if self.has_header_input.currentText() != '있음':
                    column_name = f'_c{idx}'
                else:
                    column_name = schema_item.name

                org_schema = self.schemas.get(column_name)
                if org_schema:
                    comment = org_schema.comment
                    datatype = org_schema.data_type
                    de_identi_attr = org_schema.de_identi_attr
                    is_distribution = org_schema.is_distribution
                    is_statistics = org_schema.is_statistics
                else:
                    comment = schema_item.comment
                    datatype = cnv_str_to_datatype(schema_item.data_type.name)
                    de_identi_attr = schema_item.de_identi_attr.value
                    is_distribution = schema_item.is_distribution
                    is_statistics = schema_item.is_statistics

                row_data = [
                    str(idx+1),  # 인덱스
                    column_name,
                    comment,
                    dataTypeToView(datatype),  # Enum 값 사용
                    deIdentificationAttributeToView(de_identi_attr),  # Enum 값 사용
                    is_distribution,
                    is_statistics,
                ]

                # first_rows (Pandas DataFrame)에서 해당 컬럼의 값을 가져옴
                if first_rows is not None and schema_item.name in first_rows.columns:
                    column_values_for_first_rows = first_rows[schema_item.name].tolist()
                    row_data.extend(column_values_for_first_rows)
                else:
                    row_data.extend(["N/A"] * 3)  # 3개의 샘플 값 컬럼

                # row_data의 각 요소를 QTableWidget에 추가
                row_idx = self.schema_table.rowCount()
                self.schema_table.insertRow(row_idx)

                # 데이터 타입이 숫자 타입(정수, 실수)인지 확인
                data_type_str = dataTypeToView(datatype)
                is_number_type = data_type_str in ["정수", "실수"]

                for col_idx, data_val in enumerate(row_data):
                    if col_idx in [0, 1, 7, 8, 9]:  # 순서, 컬럼명, 값1,2,3
                        item = QTableWidgetItem(str(data_val))
                        item.setTextAlignment(Qt.AlignCenter)
                        item.setFlags(item.flags() & ~Qt.ItemIsEditable)
                        self.schema_table.setItem(row_idx, col_idx, item)
                    elif col_idx == 2:  # 설명 컬럼
                        widget = QWidget()
                        layout = QHBoxLayout(widget)
                        layout.setContentsMargins(0, 0, 0, 0)
                        layout.setAlignment(Qt.AlignLeft)  # 중앙 정렬에서 왼쪽 정렬로 변경
                        line_edit = QLineEdit(str(data_val))
                        line_edit.setObjectName('comment')
                        line_edit.setStyleSheet("""
                            QLineEdit {
                                background-color: white;
                                border: 1px solid #ced4da;
                                border-radius: 4px;
                                padding: 0px 0px;
                                margin: 0px;
                                color: #495057;
                                min-height: 18px;
                                max-height: 18px;
                                height: 18px;
                                font-size: 12px;
                            }
                            QLineEdit:focus {
                                border: 1px solid #80bdff;
                                outline: none;
                            }
                            QLineEdit:disabled {
                                border: 0px;
                                background: transparent;
                            }
                        """)
                        line_edit.setAlignment(Qt.AlignLeft)  # 중앙 정렬에서 왼쪽 정렬로 변경
                        line_edit.installEventFilter(self.context_menu_filter)
                        line_edit.setDisabled(True)
                        layout.addWidget(line_edit)
                        self.schema_table.setCellWidget(row_idx, col_idx, widget)
                        self.edit_control.append(line_edit)
                    elif col_idx == 3:  # 데이터 타입
                        widget, combo = self.data_type_combo_box(data_val)
                        self.schema_table.setCellWidget(row_idx, col_idx, widget)
                        combo.setDisabled(True)
                        combo.setObjectName('data_type')
                        combo.setStyleSheet("""
                            QComboBox {
                                padding: 2px 2px;
                                border: 0px;
                                background: transparent;
                                color: #495057;
                            }
                            QComboBox::down-arrow {
                                width: 0;
                                height: 0;
                                image: none;
                                border: none;
                                background: transparent;
                            }
                            QComboBox::drop-down {
                                width: 0;
                                border: none;
                                background: transparent;
                            }
                        """)
                        combo.setEditable(True)
                        combo.lineEdit().setAlignment(Qt.AlignCenter)
                        combo.view().setItemAlignment(Qt.AlignCenter)

                        self.edit_control.append(combo)

                    elif col_idx == 4:  # 개인정보속성
                        widget, combo = self.deid_attr_combo_box(data_val)
                        self.schema_table.setCellWidget(row_idx, col_idx, widget)
                        combo.setDisabled(True)
                        combo.setObjectName('de_identi_attr')
                        combo.setStyleSheet("""
                            QComboBox {
                                padding: 2px 2px;
                                border: 0px;
                                background: transparent;
                                color: #495057;
                            }
                            QComboBox::down-arrow {
                                width: 0;
                                height: 0;
                                image: none;
                                border: none;
                                background: transparent;
                            }
                            QComboBox::drop-down {
                                width: 0;
                                border: none;
                                background: transparent;
                            }
                        """)
                        combo.setEditable(True)
                        combo.lineEdit().setAlignment(Qt.AlignCenter)
                        combo.view().setItemAlignment(Qt.AlignCenter)

                        self.edit_control.append(combo)
                    elif col_idx == 5:  # 분포: 항상 체크박스 보임
                        widget = QWidget()
                        layout = QHBoxLayout(widget)
                        layout.setContentsMargins(0, 0, 0, 0)
                        layout.setAlignment(Qt.AlignCenter)
                        checkbox = QCheckBox()
                        # checkbox.setStyleSheet(checkbox_stylesheet)
                        checkbox.setChecked(bool(data_val))
                        checkbox.setDisabled(True)
                        checkbox.setObjectName('is_distribution')
                        layout.addWidget(checkbox)
                        self.schema_table.setCellWidget(row_idx, col_idx, widget)
                        self.edit_control.append(checkbox)
                    elif col_idx == 6:  # 통계: 숫자타입만 체크박스, 아니면 숨김
                        widget = QWidget()
                        layout = QHBoxLayout(widget)
                        layout.setContentsMargins(0, 0, 0, 0)
                        layout.setAlignment(Qt.AlignCenter)
                        checkbox = QCheckBox()
                        # checkbox.setStyleSheet(checkbox_stylesheet)
                        checkbox.setChecked(bool(data_val))
                        checkbox.setDisabled(True)
                        checkbox.setObjectName('is_statistics')
                        layout.addWidget(checkbox)
                        self.schema_table.setCellWidget(row_idx, col_idx, widget)

                        if is_number_type:
                            checkbox.show()
                        else:
                            checkbox.hide()
                            # self.schema_table.setCellWidget(row_idx, col_idx, QWidget())
                            # self.edit_control.append(QWidget())
                        self.edit_control.append(checkbox)

        except Exception as e:
            self.logger.error(f"스키마 데이터 표시 중 오류 발생: {str(e)}")
            print(f"스키마 데이터 표시 중 오류 발생: {str(e)}")

    def accept(self):
        """저장 버튼 클릭 시 처리"""
        try:
            # 입력값 검증
            name = self.name_input.text().strip()  # trim 처리
            data_path = self.data_path_input.text().strip()  # trim 처리
            
            if not name:
                QMessageBox.warning(self, "입력 오류", "데이터 소스 명을 입력하세요.")
                self.name_input.setFocus()
                return
                
            if not data_path:
                QMessageBox.warning(self, "입력 오류", "원본 데이터를 선택하세요.")
                self.data_path_input.setFocus()
                return
                
            if not os.path.exists(data_path):
                QMessageBox.warning(self, "입력 오류", "입력한 데이터 경로가 존재하지 않습니다.")
                self.data_path_input.setFocus()
                return

            # 데이터 소스명 중복 체크
            for source in self.data_sources:
                if source.get('name', '').strip() == name:
                    # 수정 모드이고 자신의 이름이면 통과
                    if self.data and source.get('uuid') == self.data.get('uuid'):
                        continue
                    QMessageBox.warning(self, "입력 오류", f"'{name}' 데이터 소스명이 이미 존재합니다.")
                    self.name_input.setFocus()
                    return

            # 데이터 소스 파일 경로
            data_source_file = "datasource.json"
            temp_file = "datasource.tmp"
            backup_file = "datasource.bak"

            # 1. 임시 파일로 저장
            try:
                # 임시 파일이 이미 존재하면 삭제
                if os.path.exists(temp_file):
                    os.remove(temp_file)
                
                # 임시 파일로 저장
                with open(temp_file, 'w', encoding='utf-8') as f:
                    json.dump(self.get_data(), f, ensure_ascii=False, indent=2)
            except Exception as e:
                self.logger.error(f"임시 파일 저장 중 오류 발생: {str(e)}")
                QMessageBox.critical(self, "오류", f"데이터 소스 저장 중 오류가 발생했습니다.\n{str(e)}")
                return

            # 2. 기존 파일을 백업 파일로 이동
            try:
                if os.path.exists(data_source_file):
                    # 백업 파일이 이미 존재하면 삭제
                    if os.path.exists(backup_file):
                        os.remove(backup_file)
                    os.rename(data_source_file, backup_file)
            except Exception as e:
                self.logger.error(f"백업 파일 생성 중 오류 발생: {str(e)}")
                QMessageBox.critical(self, "오류", f"데이터 소스 백업 중 오류가 발생했습니다.\n{str(e)}")
                return

            # 3. 임시 파일을 실제 파일로 이동
            try:
                os.rename(temp_file, data_source_file)
            except Exception as e:
                self.logger.error(f"파일 이동 중 오류 발생: {str(e)}")
                # 오류 발생 시 백업 파일을 원래대로 복구
                if os.path.exists(backup_file):
                    if os.path.exists(data_source_file):
                        os.remove(data_source_file)
                    os.rename(backup_file, data_source_file)
                QMessageBox.critical(self, "오류", f"데이터 소스 저장 중 오류가 발생했습니다.\n{str(e)}")
                return

            # 4. 백업 파일 삭제
            try:
                if os.path.exists(backup_file):
                    os.remove(backup_file)
            except Exception as e:
                self.logger.error(f"백업 파일 삭제 중 오류 발생: {str(e)}")
                # 백업 파일 삭제 실패는 치명적이지 않으므로 경고만 표시
                QMessageBox.warning(self, "경고", f"백업 파일 삭제 중 오류가 발생했습니다.\n{str(e)}")
            
            # 모든 검증을 통과하면 부모 클래스의 accept() 호출
            super().accept()
            
        except Exception as e:
            self.logger.error(f"데이터 소스 저장 중 오류 발생: {str(e)}")
            QMessageBox.critical(self, "오류", f"데이터 소스 저장 중 오류가 발생했습니다.\n{str(e)}")

    def get_data(self):
        data_path = self.data_path_input.text()
        if data_path == '' or not os.path.exists(data_path):
            self.logger.error(f"데이터 소스 경로가 존재하지 않습니다. : {data_path}")
            return None
        
        data_type = get_file_extension(data_path.lower())

        if os.path.isdir(data_path):
            type = 'dir'
        else:
            type = 'file'

        schemas = []
        for row in range(self.schema_table.rowCount()):
            schema = Schema()
            
            # 컬럼명
            name_item = self.schema_table.item(row, 1)
            schema.name = name_item.text() if name_item else f'_c{row}'
            
            # 설명
            comment_widget = self.schema_table.cellWidget(row, 2)
            comment_line_edit = comment_widget.findChild(QLineEdit) if comment_widget else None
            schema.comment = comment_line_edit.text() if comment_line_edit else ''
            
            # 데이터 타입
            data_type_widget = self.schema_table.cellWidget(row, 3)
            data_type_combobox = data_type_widget.findChild(QComboBox) if data_type_widget else None
            if data_type_combobox is not None:
                data_type_item = data_type_combobox.currentText() if data_type_combobox else None
            else:
                data_type_item = "문자"
            schema.data_type = viewToDataType(data_type_item) if data_type_item else DataType.STRING
            
            # 개인정보속성
            deid_widget = self.schema_table.cellWidget(row, 4)
            deid_combobox = deid_widget.findChild(QComboBox) if deid_widget else None
            if deid_combobox is not None:
                deid_item = deid_combobox.currentText() if deid_combobox else None
            else:
                deid_item = "문자"
            schema.de_identi_attr = viewToDeIdentificationAttribute(deid_item) if deid_item else DeIdentificationAttribute.NONE
            
            # 분포
            dist_widget = self.schema_table.cellWidget(row, 5)
            dist_checkbox = dist_widget.findChild(QCheckBox) if dist_widget else None
            if dist_checkbox is not None:
                schema.is_distribution = dist_checkbox.isChecked()
            else:
                schema.is_distribution = False
            # 통계
            stats_widget = self.schema_table.cellWidget(row, 6)
            stats_checkbox = stats_widget.findChild(QCheckBox) if stats_widget else None
            if stats_checkbox is not None:
                if schema.data_type in [DataType.INTEGER, DataType.LONG, DataType.FLOAT, DataType.DECIMAL, DataType.DOUBLE]:
                    schema.is_statistics = stats_checkbox.isChecked()
                else:
                    schema.is_statistics = False
            else:
                schema.is_statistics = False

            schema.exception_text = ''
            schema.json = ''
            
            # Schema 객체를 딕셔너리로 변환
            schema_dict = {
                'name': schema.name,
                'comment': schema.comment,
                'data_type': schema.data_type.name,
                'de_identi_attr': schema.de_identi_attr.name,
                'is_distribution': schema.is_distribution,
                'is_statistics': schema.is_statistics,
                'exception_text': schema.exception_text,
                'json': schema.json
            }
            schemas.append(schema_dict)

        return {
            'uuid': str(self.uuid),  # UUID를 문자열로 변환
            'name': self.name_input.text(),
            'type': type,
            'data_type': data_type,
            'data_path': self.data_path_input.text(),
            'charset': self.charset_input.currentText(),
            'separator': self.separator_input.currentText(),
            'has_header': self.has_header_input.currentText(),
            'schemas': schemas
        }

    def get_schema(self, df):
        """스키마 정보 가져오기"""
        dtypes = df.dtypes
        
        # 컬럼명과 데이터 타입 출력
        schrmas = []
        for column_name, data_type in dtypes.items():
            # row = [column, dtype]
            schema = Schema()
            schema.name = column_name
            schema.data_type = data_type
            schrmas.append(schema)
        
        return schrmas

    def _add_check_column(self, is_checked: bool):
        checkbox_item = QStandardItem()
        checkbox_item.setCheckable(True) # 체크박스 활성화
        checkbox_item.setEditable(True) # 체크박스 텍스트는 편집 불가능하게
        checkbox_item.setCheckState(Qt.Checked if is_checked else Qt.UnChecked) # 초기 상태는 체크 해제

        return checkbox_item

    def browse_data_path(self):
        """데이터 파일/폴더 선택 (QMenu 방식)"""
        menu = QMenu(self)
        menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #dee2e6;
                border-radius: 4px;
                padding: 5px;
            }
            QMenu::item {
                padding: 8px 20px;
                color: #495057;
            }
            QMenu::item:selected {
                background-color: #e9ecef;
                color: #000;
            }
        """)
        file_action = QAction("파일 선택", self)
        dir_action = QAction("폴더 선택", self)
        menu.addAction(file_action)
        menu.addAction(dir_action)

        def select_file():
            start_path = self.data_path_input.text().strip()
            if not start_path or not os.path.exists(start_path):
                start_path = os.path.expanduser('~')
            file_path, _ = QFileDialog.getOpenFileName(
                self, "원본 데이터 파일 선택", start_path,
                "CSV 파일 (*.csv);;.parquet 파일 (*.parquet);;모든 파일 (*.*)"
            )
            if file_path:
                if platform.system() == 'Windows':
                    file_path = file_path.replace('/', '\\')
                self.data_path_input.setText(file_path)
                
                # CSV 파일인 경우 파일을 읽어서 컬럼명과 데이터 타입 출력
                # if file_path.lower().endswith('.csv'):
                try:
                    self.is_csv = True if file_path.lower().endswith('.csv') else False

                    if self.is_csv:
                        self.csv_container.show()
    
                        # 파일의 인코딩 감지
                        with open(file_path, 'rb') as file:
                            raw_data = file.read()
                            result = chardet.detect(raw_data)
                            detected_encoding = result['encoding']
                            confidence = result['confidence']
                            
                            # UTF-8-SIG(BOM)을 UTF-8로 변환
                            if detected_encoding and detected_encoding.upper() == 'UTF-8-SIG':
                                detected_encoding = 'UTF-8'
                            
                            print(f"감지된 파일 인코딩: {detected_encoding} (신뢰도: {confidence:.2%})")
                            
                            # 감지된 인코딩이 있으면 콤보박스에 설정
                            if detected_encoding:
                                index = self.charset_input.findText(detected_encoding.upper())
                                if index >= 0:
                                    self.charset_input.setCurrentIndex(index)
                    else:
                        self.csv_container.hide()

                    
                    # CSV 파일 읽기 및 테이블 업데이트
                    self.message_container.hide()
                    self.load_data()
                except Exception as e:
                    self.logger.error(f"CSV 파일 읽기 중 오류 발생: {str(e)}")
                    self.message_text.setText("CSV 파일 읽기 중 오류가 발생했습니다. 원본 데이터를 확인해주세요.")
                    self.message_container.show()

        def select_dir():
            start_path = self.data_path_input.text().strip()
            if not start_path or not os.path.exists(start_path):
                start_path = os.path.expanduser('~')
            dir_path = QFileDialog.getExistingDirectory(
                self, "원본 데이터 폴더 선택", start_path
            )
            if dir_path:
                if platform.system() == 'Windows':
                    dir_path = dir_path.replace('/', '\\')
                self.data_path_input.setText(dir_path)
                try:
                    self.csv_container.hide()
                
                    # CSV 파일 읽기 및 테이블 업데이트
                    self.message_container.hide()
                    self.load_data()
                except Exception as e:
                    self.logger.error(f"데이터 읽기 중 오류 발생: {str(e)}")
                    self.message_text.setText("데이터 읽기 중 오류가 발생했습니다. 원본 데이터를 확인해주세요.")
                    self.message_container.show()

        file_action.triggered.connect(select_file)
        dir_action.triggered.connect(select_dir)
        menu.exec_(self.fs_explorer.mapToGlobal(self.fs_explorer.rect().bottomLeft())) 

    def load_data(self):
        """인코딩이 변경될 때 CSV 파일을 다시 읽어서 테이블 업데이트"""
        try:
            self.show_loading()
            file_path = self.data_path_input.text()
            if not file_path:
                return 
            self.is_csv = True if file_path.lower().endswith('.csv') else False

            if self.is_csv:
                header_line = 0 if self.has_header_input.currentText() == '있음' else None
                
                sep = self.separator_input.currentText()
                if sep == 'TAB':
                    sep = '\t'
                self.csv_container.show()
            elif file_path.lower().endswith('.parquet'):
                self.csv_container.hide()
            else:
                raise ValueError(f"지원하지 않는 파일 형식입니다. : {file_path}")

            if os.path.exists(file_path):
                # dask를 사용하여 CSV 파일 읽기
                if self.is_csv:
                    df = read_csv(file_path, 
                                encoding=self.charset_input.currentText(),
                                sep=sep,
                                header=header_line,
                                encoding_errors='replace')
                else: # file_path.lower().endswith('.parquet'):
                    df = read_parquet(file_path)

                file_schemas = self.get_schema(df)
                first_rows = df.head(3)

                # 스키마 데이터 표시
                self.display_schema_data(file_schemas, first_rows)
            else:
                raise ValueError(f"선택한 데이타가 존재하지 않습니다. : {file_path}")
                
                
                
        except Exception as e:
            self.logger.error(f"원본 데이터 읽기 중 오류 발생: {str(e)}")
            self.message_text.setText("원본 데이터 읽기 중 오류가 발생했습니다. 원본 데이터를 확인해주세요.")
            self.message_container.show()

            # print(f"CSV 파일 다시 읽기 중 오류 발생: {str(e)}")
        finally:
            self.adjust_widget_sizes()
            self.hide_loading()

    def show_loading(self):
        if self.loading_overlay:
            self.loading_overlay.show()
            self.loading_movie.start()
            QApplication.processEvents()  # UI 즉시 갱신

    def hide_loading(self):
        if self.loading_overlay:
            self.loading_overlay.hide()
            self.loading_movie.stop()

    def eventFilter(self, obj, event):
        """이벤트 필터"""
        if obj == self and event.type() == QEvent.Type.Resize:
            # 다이얼로그 크기가 변경될 때 내부 위젯 크기 조절
            self.adjust_widget_sizes()
        return super().eventFilter(obj, event)

    def adjust_widget_sizes(self):
        """내부 위젯들의 크기를 조절합니다."""
        try:
            # 다이얼로그의 현재 크기
            dialog_width = self.width()
            dialog_height = self.height()
            
            # 스키마 테이블 크기 조절
            if hasattr(self, 'schema_table'):
                # 다이얼로그 높이에 따라 테이블 높이 조절
                    # 상단 영역(입력 필드들), 하단 영역(버튼들), form_container 마진(24px)을 제외한 공간 계산
                available_height = 167 + dialog_height - 627 # 상단 200px + 하단 100px + 마진 24px
                available_width = 730  + dialog_width - 800

                file_path = self.data_path_input.text()
                if self.csv_container.isHidden():
                    available_height += 144
                if self.message_container.isHidden():
                    available_height += 26 #21

                # 테이블 높이를 다이얼로그 높이의 60%로 제한
                # max_height = int(dialog_height * 0.6)
                self.schema_table.setFixedHeight(max(161, available_height))
                
                # 테이블 너비는 항상 730px 유지
                self.schema_table.setFixedWidth(max(730, available_width))
                
        except Exception as e:
            self.logger.error(f"위젯 크기 조절 중 오류 발생: {str(e)}") 

    
    def edit_schema(self):
        self.change_combobox_style(self.edit_flag)

    def change_combobox_style(self, edit_flag):
        """스키마 편집 상태로"""

        for control in self.edit_control:
            if not self.is_csv and control.objectName() == 'data_type':
                control.setDisabled(True)
            else:
                control.setDisabled(edit_flag)

            if isinstance(control, QComboBox) and \
                (self.is_csv or not self.is_csv and control.objectName() != 'data_type'):

                if not self.edit_flag:
                    control.setStyleSheet("""
                        QComboBox {
                            background-color: white;
                            border: 1px solid #dee2e6;
                            border-radius: 4px;
                            padding: 2px 2px;
                            color: #495057;
                        }
                        QComboBox::drop-down {
                            subcontrol-origin: padding;
                            subcontrol-position: center right;
                            width: 30px;
                            border: none;
                        }
                        QComboBox::down-arrow {
                            image: url(src/ui/resources/images/down-arrow-svgrepo-com.svg);
                            width: 10px;
                            height: 10px;
                            margin-right: 1px;
                        }
                        QComboBox QAbstractItemView {
                            background-color: white;
                            border: 1px solid #dee2e6;
                            border-radius: 4px;
                            outline: none;
                            selection-background-color: #e9ecef;
                            selection-color: #000;
                        }
                        QComboBox QAbstractItemView::item {
                            padding: 2px 2px;
                            color: #495057;
                        }
                        QComboBox QAbstractItemView::item:selected {
                            background-color: #e9ecef;
                            color: #000;
                        }
                    """)
                    control.setEditable(False)
                else:
                    control.setStyleSheet("""
                        QComboBox {
                            padding: 2px 2px;
                            border: 0px;
                            background: transparent;
                            color: #495057;
                        }
                        QComboBox::down-arrow {
                            width: 0;
                            height: 0;
                            image: none;
                            border: none;
                            background: transparent;
                        }
                        QComboBox::drop-down {
                            width: 0;
                            border: none;
                            background: transparent;
                        }
                    """)
                    control.setEditable(True)
                    control.lineEdit().setAlignment(Qt.AlignCenter)
                    control.view().setItemAlignment(Qt.AlignCenter)                    
        self.edit_flag = not edit_flag

        if self.edit_flag:
            self.schema_edit_btn.setIcon(QIcon("src/ui/resources/images/icons8-unlock-96.png"))
            # self.schema_edit_btn.setIconSize(QSize(16, 16))
            # self.schema_edit_btn.setText("")
            self.schema_edit_btn.setToolTip("스키마 편집")
        else:
            self.schema_edit_btn.setIcon(QIcon("src/ui/resources/images/icons8-lock-96.png"))
            # self.schema_edit_btn.setIconSize(QSize(16, 16))
            # self.schema_edit_btn.setText("")
            self.schema_edit_btn.setToolTip("스키마 편집 잠김")
            # if isinstance(control, QLineEdit):
            #     control.setDisabled(False)
            # elif isinstance(control, QComboBox):
            #     control.setDisabled(False)
            # elif isinstance(control, QCheckBox):
            #     control.setDisabled(False)

    def download_schema_excel(self):
        try:
            template_path = "templates/table_info.xlsx"
            if not os.path.exists(template_path):
                QMessageBox.critical(self, "오류", "엑셀 템플릿 파일이 존재하지 않습니다.")
                return

            wb = openpyxl.load_workbook(template_path)
            ws = wb.active

            # 1. 테이블 정보 추출 및 매핑
            columns = []
            from defined import DeIdentificationAttribute
            for row in range(self.schema_table.rowCount()):
                # 컬럼번호, 컬럼명, 설명, 분포, 통계, 데이터타입, 개인정보속성
                col_num = str(row + 1)
                col_name = self.schema_table.item(row, 1).text() if self.schema_table.item(row, 1) else ""
                cell_widget = self.schema_table.cellWidget(row, 2)
                if cell_widget:
                    line_edit = cell_widget.findChild(QLineEdit)
                    if line_edit:
                        comment = line_edit.text()  # QLineEdit의 텍스트 값 가져오기
                    else:
                        comment = ""
                else:
                    comment = ""                
                # 분포
                dist_widget = self.schema_table.cellWidget(row, 5)
                dist_checkbox = dist_widget.findChild(QCheckBox) if dist_widget else None
                if dist_checkbox is not None:
                    dist_val = '○' if dist_checkbox.isChecked() else ''
                else:
                    dist_val = ''
                # 통계
                stats_widget = self.schema_table.cellWidget(row, 6)
                stats_checkbox = stats_widget.findChild(QCheckBox) if stats_widget else None
                if stats_checkbox is not None:
                    stats_val = '○' if stats_checkbox.isChecked() else ''
                else:
                    stats_val = ''
                # 데이터타입
                cell_widget = self.schema_table.cellWidget(row, 3)
                if cell_widget:
                    combo = cell_widget.findChild(QComboBox)
                    if combo:
                        data_type = combo.currentText()  # QLineEdit의 텍스트 값 가져오기
                    else:
                        data_type = ""
                else:
                    data_type = ""
                # 개인정보속성 (value)
                cell_widget = self.schema_table.cellWidget(row, 4)
                if cell_widget:
                    line_edit = cell_widget.findChild(QComboBox)
                    if line_edit:
                        de_identi_attr = line_edit.currentText()  # QLineEdit의 텍스트 값 가져오기
                    else:
                        de_identi_attr = ""
                else:
                    de_identi_attr = ""
                try:
                    de_identi_value = DeIdentificationAttribute[de_identi_attr].value
                except Exception:
                    de_identi_value = de_identi_attr

                columns.append([
                    col_num, col_name, comment, dist_val, stats_val, data_type, de_identi_value
                ])

            # 2. A3부터 데이터 입력
            start_row = 3
            # 헤더 폰트 추출 (A2~G2가 헤더라고 가정)
            header_font = ws['A2'].font.copy()
            from openpyxl.styles import Border, Side, Alignment
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for idx, col_data in enumerate(columns):
                for j, value in enumerate(col_data):
                    cell = ws.cell(row=start_row + idx, column=j + 1, value=value)
                    # 중앙정렬 대상 컬럼: 0(번호), 3(분포), 4(통계), 5(데이터타입), 6(개인정보속성)
                    if j in [0, 3, 4, 5, 6]:
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                    # 폰트 헤더와 동일하게
                    cell.font = header_font
                    # 테두리
                    cell.border = border

            # 4. 저장 경로 생성 및 저장 (이전 코드와 동일)
            datasource_name = self.name_input.text().strip()
            if not datasource_name:
                QMessageBox.warning(self, "입력 오류", "데이터 소스 명을 입력하세요.")
                return

            data_directory = None
            parent = self.parent()
            if parent and hasattr(parent, 'project_data'):
                data_directory = parent.project_data.get('data_directory')
            if not data_directory:
                QMessageBox.critical(self, "오류", "프로젝트의 data_directory 정보를 찾을 수 없습니다.")
                return

            table_info_dir = os.path.join(data_directory, "01.테이블 정보")
            os.makedirs(table_info_dir, exist_ok=True)

            save_path = os.path.join(table_info_dir, f"{datasource_name}_테이블_정의.xlsx")
            try:
                wb.save(save_path)
                QMessageBox.information(self, "완료", f"테이블 정의서가 생성되었습니다:\n{save_path}")
            except PermissionError:
                QMessageBox.critical(self, "오류", "테이블 정의서를 생성 할 수 없습니다. 이미 열려 있다면 먼저 닫아 주세요")
            except Exception as e:
                self.logger.error(f"테이블 정의서 생성 중 오류: {str(e)}")
                QMessageBox.critical(self, "오류", f"테이블 정의서 생성 중 오류가 발생했습니다.\n{str(e)}")

        except Exception as e:
            self.logger.error(f"테이블 정의서 생성 중 오류: {str(e)}")
            QMessageBox.critical(self, "오류", f"테이블 정의서 생성 중 오류가 발생했습니다.\n{str(e)}")

    def upload_schema_excel(self):
        from PySide6.QtWidgets import QFileDialog
        # 1. 파일 탐색기 (엑셀 파일만)
        file_path, _ = QFileDialog.getOpenFileName(
            self, "테이블 정의서 엑셀 파일 선택", "",
            "Excel Files (*.xlsx *.xls)"
        )
        if not file_path:
            return

        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active

            # 2. 테이블 정의서 포맷 확인 (A2~G2에 헤더가 있어야 함)
            expected_headers = ["컬럼번호", "컬럼명", "설명", "분포", "통계", "데이터타입", "개인정보속성"]
            headers = [ws.cell(row=2, column=i+1).value for i in range(len(expected_headers))]
            if headers != expected_headers:
                QMessageBox.critical(self, "오류", "테이블정의 파일이 아닙니다.")
                return

            # 3. 데이터 읽기 (A3부터)
            data_list = []
            row_idx = 3
            while True:
                row = [ws.cell(row=row_idx, column=col+1).value for col in range(len(expected_headers))]
                # 컬럼명 없으면 종료
                if not row[1]:
                    break
                # 분포/통계 변환
                row[3] = True if row[3] == '○' else False
                row[4] = True if row[4] == '○' else False
                # 설명/개인정보속성 None 처리
                if row[2] is None:
                    row[2] = ""
                if row[6] is None:
                    row[6] = ""
                data_list.append(row)
                row_idx += 1

            # 4. 기존 테이블 구조는 그대로 두고 값만 변경
            # 4-1. 컬럼명 일치 여부 확인
            excel_colnames = [str(row[1]).strip() if row[1] is not None else "" for row in data_list]
            ui_colnames = []
            for i in range(self.schema_table.rowCount()):
                item = self.schema_table.item(i, 1)
                ui_colnames.append(item.text().strip() if item and item.text() else "")
            if len(excel_colnames) != len(ui_colnames) or excel_colnames != ui_colnames:
                QMessageBox.critical(self, "오류", "원본데이타의 컬럼과 테이블 정의서의 컬럼이 다릅니다.")
                return

            row_count = self.schema_table.rowCount()
            for i, row in enumerate(data_list):
                if i >= row_count:
                    break  # 기존 테이블보다 엑셀 데이터가 많으면 무시
                # 데이터타입
                data_type = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                is_number_type = data_type in ["정수", "실수"]
                # 통계: 숫자 타입이 아니면 무조건 False
                if not is_number_type:
                    row[4] = False
                # 컬럼명
                item = self.schema_table.item(i, 1)
                if item:
                    item.setText(str(row[1]))
                # 설명
                item = self.schema_table.cellWidget(i, 2)
                if item:
                    line_edit = item.findChild(QLineEdit)
                    if line_edit:
                        line_edit.setText(str(row[2])) 
                # 데이터타입
                item = self.schema_table.cellWidget(i, 3)
                if item:
                    combo = item.findChild(QComboBox)
                    if combo:
                        combo.setCurrentText(str(row[5]))                     
                # 개인정보속성
                item = self.schema_table.cellWidget(i, 4)
                if item:
                    combo = item.findChild(QComboBox)
                    if combo:
                        combo.setCurrentText(str(row[6]))                     
                # 분포(체크박스)
                widget = self.schema_table.cellWidget(i, 5)
                if widget:
                    checkbox = widget.findChild(QCheckBox)
                    if checkbox:
                        checkbox.setChecked(bool(row[3]))
                # 통계(체크박스)
                widget = self.schema_table.cellWidget(i, 6)
                checkbox = widget.findChild(QCheckBox) if widget else None
                if is_number_type:
                    # 숫자 타입인데 체크박스가 없으면 새로 생성
                    if checkbox is None:
                        new_widget = QWidget()
                        layout = QHBoxLayout(new_widget)
                        layout.setContentsMargins(0, 0, 0, 0)
                        layout.setAlignment(Qt.AlignCenter)
                        checkbox = QCheckBox()
                        layout.addWidget(checkbox)
                        new_widget.setLayout(layout)
                        self.schema_table.setCellWidget(i, 6, new_widget)
                    checkbox.setChecked(bool(row[4]))
                    checkbox.setVisible(True)
                    checkbox.setDisabled(True)
                else:
                    # 숫자 타입이 아니면 체크박스가 있으면 unchecked + 숨김
                    if checkbox is not None:
                        checkbox.setChecked(False)
                        checkbox.setVisible(False)

        except Exception as e:
            self.logger.error(f"테이블 정의서 업로드 오류: {str(e)}")
            QMessageBox.critical(self, "오류", f"테이블 정의서 업로드 중 오류가 발생했습니다.\n{str(e)}")

